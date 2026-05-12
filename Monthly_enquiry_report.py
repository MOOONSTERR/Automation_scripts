import os
import re
import logging
import base64
import requests
import msal
import pytz
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================== 多租户配置区域 ==================
MAILBOX_CONFIGS = [
    {
        "email": "sample_1@email.com",
        "tenant_id": "sample_tenant_id_1",
        "client_id": "sample_client_id_1",
        "client_secret": "sample_secret_1"
    },
    {
        "email": "sample_2@email.com",
        "tenant_id": "sample_tenant_id_2",
        "client_id": "sample_client_id_2",
        "client_secret": "sample_secret_2"
    }
]

SENDER_EMAIL = "sample_1@email.com"
REPORT_RECIPIENT = "hubin@jpnelson.com.sg"
CC_RECIPIENT = "hubin@jpnelson.com.sg"

TARGET_RECIPIENTS = {
    'sample_1@email.com', 'sample_2@email.com', 'sample_3@email.com',
    'User 1', 'User 2', 'User 3'
}
# ===================================================

SG_TZ = pytz.timezone('Asia/Singapore')

def setup_logging():
    logs_dir = os.path.join("Logs", "Monthly_Report_Log")
    os.makedirs(logs_dir, exist_ok=True)
    date_str = datetime.now(SG_TZ).strftime('%d_%m_%Y')
    log_file = os.path.join(logs_dir, f"Monthly_Enquiry_Summary_log_{date_str}.log")
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - [%(levelname)s] - %(message)s',
        handlers=[logging.FileHandler(log_file, encoding='utf-8')]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

def normalize_subject(subject):
    if not subject: return ""
    sub = re.sub(r'^(FW|FWD|RE|答复|转发|External|Automatic reply):\s*', '', subject.strip(), flags=re.IGNORECASE)
    sub = re.sub(r'\[.*?\]', '', sub)
    return ' '.join(sub.lower().split()).strip()

def get_token(config):
    authority = f"https://login.microsoftonline.com/{config['tenant_id']}"
    app = msal.ConfidentialClientApplication(config['client_id'], authority=authority, client_credential=config['client_secret'])
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return result.get("access_token")

def fetch_messages_cache(token, mailbox, start_date, end_date):
    cache = {}
    url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages"
    headers = {"Authorization": f"Bearer {token}"}
    start_utc = start_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    end_utc = end_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    params = {
        "$select": "subject,receivedDateTime,sender,toRecipients",
        "$filter": f"receivedDateTime ge {start_utc} and receivedDateTime le {end_utc}",
        "$orderby": "receivedDateTime desc", "$top": 999
    }
    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code != 200: break
        data = resp.json()
        for item in data.get('value', []):
            clean_sub = normalize_subject(item.get('subject', ''))
            sender = (item.get('sender') or {}).get('emailAddress', {}).get('address', '')
            to_list = [r.get('emailAddress', {}).get('address', '') for r in item.get('toRecipients', [])]
            dt = datetime.strptime(item['receivedDateTime'][:19], '%Y-%m-%dT%H:%M:%S').replace(tzinfo=pytz.utc).astimezone(SG_TZ)
            cache[clean_sub] = {
                'received_time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'sender': sender, 'to': "; ".join(filter(None, to_list)),
                'full_subject': item.get('subject', '')
            }
        url = data.get('@odata.nextLink'); params = None
    return cache

def fetch_sent_forwards(token, mailbox, start_date, end_date):
    forwards = []
    url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/mailFolders/SentItems/messages"
    headers = {"Authorization": f"Bearer {token}"}
    start_utc = start_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    end_utc = end_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    params = {
        "$select": "subject,sentDateTime,toRecipients",
        "$filter": f"sentDateTime ge {start_utc} and sentDateTime le {end_utc}",
        "$orderby": "sentDateTime desc"
    }
    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code != 200: break
        data = resp.json()
        for item in data.get('value', []):
            subject = item.get('subject', '')
            if not any(p in subject.upper() for p in ['FW:', 'FWD:']): continue
            matches = [r['emailAddress'].get('name') or r['emailAddress'].get('address') for r in item.get('toRecipients', []) 
                       if any(t.lower() in (r['emailAddress'].get('name','') + r['emailAddress'].get('address','')).lower() for t in TARGET_RECIPIENTS)]
            if matches:
                dt = datetime.strptime(item['sentDateTime'][:19], '%Y-%m-%dT%H:%M:%S').replace(tzinfo=pytz.utc).astimezone(SG_TZ)
                forwards.append({
                    'raw_subject': subject, 'clean_subject': normalize_subject(subject),
                    'forwarded_to': ", ".join(sorted(set(matches))),
                    'forwarded_on': dt.strftime('%Y-%m-%d %H:%M:%S'), 'sort_key': dt
                })
        url = data.get('@odata.nextLink'); params = None
    return forwards

def main():
    logger.info("--- Monthly Report Started ---")
    now_sg = datetime.now(SG_TZ)
    # 日期定义
    first_day_this_month = now_sg.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_last_month = first_day_this_month - timedelta(seconds=1)
    first_day_last_month = (first_day_this_month - timedelta(days=28)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_day_history = (first_day_last_month - timedelta(days=5)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    combined_data = []
    sender_token = None

    for config in MAILBOX_CONFIGS:
        token = get_token(config)
        if not token: continue
        if config['email'] == SENDER_EMAIL: sender_token = token
        
        msg_cache = fetch_messages_cache(token, config['email'], first_day_history, now_sg)
        forwards = fetch_sent_forwards(token, config['email'], first_day_last_month, last_day_last_month)
        
        for f in forwards:
            match = msg_cache.get(f['clean_subject'])
            if not match:
                for k, v in msg_cache.items():
                    if f['clean_subject'] and (f['clean_subject'] in k or k in f['clean_subject']):
                        match = v; break
            combined_data.append({
                'Source Mailbox': config['email'],
                'Sent on': match['received_time'] if match else '',
                'Sender Address': match['sender'] if match else '',
                'Sent to': match['to'] if match else '',
                'Forwarded to': f['forwarded_to'],
                'Forwarded on': f['forwarded_on'],
                'Title': match['full_subject'] if match else f['raw_subject'],
                '_sort_date': f['sort_key']
            })

    if not combined_data:
        logger.warning("No records found."); return

    df = pd.DataFrame(combined_data).sort_values(by='_sort_date', ascending=False).drop(columns=['_sort_date'])

    # 修改命名的核心代码
    # %b 会生成英文月份简写 (Jan, Feb, Mar...)
    report_date_str = first_day_last_month.strftime('%b_%Y') 
    report_name = f"Monthly_Enquiry_Summary_{report_date_str}.xlsx"
    
    output_dir = "Monthly_Reports"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, report_name)
    
    df.to_excel(output_path, index=False)
    
    # 格式化
    wb = load_workbook(output_path); ws = wb.active
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True); cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    wb.save(output_path)

    # 发送邮件
    if sender_token:
        try:
            with open(output_path, "rb") as f: content = base64.b64encode(f.read()).decode()
            payload = {
                "message": {
                    "subject": f"Monthly Enquiry Summary {first_day_last_month.strftime('%B %Y')}",
                    "body": {"contentType": "Text", "content": f"Hi Mr Nelson,\n\nPlease find the attached monthly report for {first_day_last_month.strftime('%B %Y')}, thanks."},
                    "toRecipients": [{"emailAddress": {"address": REPORT_RECIPIENT}}],
                    "ccRecipients": [{"emailAddress": {"address": REPORT_RECIPIENT}}],
                    "attachments": [{"@odata.type": "#microsoft.graph.fileAttachment", "name": report_name, "contentBytes": content}]
                }
            }
            requests.post(f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail", headers={"Authorization": f"Bearer {sender_token}"}, json=payload)
            logger.info("Success: Email sent.")
        except Exception as e: logger.error(f"Error: {e}")

if __name__ == "__main__":
    main()