import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.colors import Color
from openpyxl.worksheet.table import Table, TableStyleInfo # 导入Table相关模块
from openpyxl.chart import BarChart, Reference # 导入图表相关模块
import os
from openpyxl.utils import get_column_letter # 导入获取列字母的函数

# --- 配置 ---
DATA_SOURCES_INV_CN = {
    'inv_cn': ['RY.xlsx', 'CNR.xlsx', 'sales-inv.xlsx', 'sales-cn.xlsx'],
}
DATA_SOURCES_LINE = {
    'line': ['inv-line.xlsx', 'cnr-line.xlsx'],
}
COA_FILE = 'COA.xlsx'
TEMPLATE_FILE = 'Template.xlsx'  # 修改了这里

# COA 到 Final Working 的列映射
COA_TO_FINAL_WORKING_MAP = {
    'Posting Date': 'Posting Date',
    'Document Type': 'Document Type',
    'Document No.': 'Document No.',
    'G/L Account No.': 'G/L Account No.',
    'G/L Account Name': 'Account Name',
    'Debit Amount': 'Debit Amount',
    'Credit Amount': 'Credit Amount',
}

print("开始执行完整集成流程...")

# --- 辅助函数 ---
def load_and_preprocess_inv_cn(files):
    df_parts = []
    for file in files:
        try:
            df_temp = pd.read_excel(file)
            print(f"    - 成功加载 {file}, 形状: {df_temp.shape}")

            if file in ['RY.xlsx', 'CNR.xlsx']:
                if 'Name' in df_temp.columns:
                    df_temp.rename(columns={'Name': 'Customer Name'}, inplace=True)
                    print(f"      - 已将 '{file}' 中的 'Name' 列重命名为 'Customer Name'。")
            elif file == 'sales-cn.xlsx':
                if 'Sell-to Customer No.' in df_temp.columns:
                    df_temp.rename(columns={'Sell-to Customer No.': 'Customer No.'}, inplace=True)
                    print(f"      - 已将 '{file}' 中的 'Sell-to Customer No.' 列重命名为 'Customer No.'。")
            
            df_parts.append(df_temp)

        except FileNotFoundError:
            print(f"    - 警告: 文件 {file} 未找到，跳过。")
        except Exception as e:
            print(f"    - 读取 {file} 时出错: {e}")
    
    return pd.concat(df_parts, ignore_index=True) if df_parts else pd.DataFrame()

def load_and_preprocess_line(files):
    df_parts = []
    for file in files:
        try:
            df_temp = pd.read_excel(file)
            print(f"    - 成功加载 {file}, 形状: {df_temp.shape}")
            df_parts.append(df_temp)
        except FileNotFoundError:
            print(f"    - 警告: 文件 {file} 未找到，跳过。")
        except Exception as e:
            print(f"    - 读取 {file} 时出错: {e}")
    return pd.concat(df_parts, ignore_index=True) if df_parts else pd.DataFrame()

def preprocess_coa(df):
    if df.empty:
        return df
        
    print("    - 开始预处理 COA 数据...")
    
    # --- 步骤 1: 删除最后一行 ---
    print(f"      - 预处理前 COA 形状: {df.shape}")
    df = df.iloc[:-1]  # 删除最后一行
    print(f"      - 删除最后一行后，COA 形状: {df.shape}")

    df['Debit Amount'] = pd.to_numeric(df['Debit Amount'], errors='coerce')
    df['Credit Amount'] = pd.to_numeric(df['Credit Amount'], errors='coerce')

    initial_rows = len(df)
    df = df[~((df['Debit Amount'] == 0) & (df['Credit Amount'] == 0))]
    final_rows = len(df)
    print(f"      - 删除了 {initial_rows - final_rows} 行全零记录。")

    df = df.replace(0, pd.NA)
    print("      - 已将所有0值替换为 NA。")

    if 'Debit Amount' in df.columns:
        df['Debit Amount'] = df['Debit Amount'].abs() * -1
        print("      - 已将 Debit Amount 转换为负数。")
    
    print(f"    - 预处理后 COA 形状: {df.shape}")
    return df

# --- 步骤 1: 处理 INV & CN 数据 ---
print("\n--- 步骤 1: 加载并预处理 INV & CN 数据源 ---")
df_inv_cn = load_and_preprocess_inv_cn(DATA_SOURCES_INV_CN['inv_cn'])
print(f"INV & CN 合并后形状: {df_inv_cn.shape}")

# --- 步骤 2: 处理 LINE 数据 ---
print("\n--- 步骤 2: 加载并合并 LINE 数据源 ---")
df_line = load_and_preprocess_line(DATA_SOURCES_LINE['line'])
print(f"LINE 合并后形状: {df_line.shape}")

# --- 步骤 3: 加载 Cust-Indust-M 数据 ---
print("\n--- 步骤 3: 加载 Cust-Indust-M 数据 ---")
try:
    df_cust_industry = pd.read_excel(TEMPLATE_FILE, sheet_name='Cust-Indust-M')
    print(f"    - 成功加载 {TEMPLATE_FILE} 中的 'Cust-Indust-M' 表, 形状: {df_cust_industry.shape}")
    print(f"    - Cust-Indust-M 列: {list(df_cust_industry.columns)}")
except Exception as e:
    print(f"    - 读取 'Cust-Indust-M' 表时出错: {e}")
    df_cust_industry = pd.DataFrame()

# --- 步骤 4: 预处理 COA 数据并映射 ---
print("\n--- 步骤 4: 预处理 COA 数据并映射到 Final Working 结构 ---")

try:
    df_coa_raw = pd.read_excel(COA_FILE)
    print(f"    - 成功加载 {COA_FILE}, 形状: {df_coa_raw.shape}")
    
    # --- 新增步骤: 提取月份和年份 ---
    if 'Posting Date' in df_coa_raw.columns:
        posting_dates = pd.to_datetime(df_coa_raw['Posting Date'], format='%d/%m/%Y', errors='coerce')
        valid_dates = posting_dates.dropna()
        if not valid_dates.empty:
            first_date = valid_dates.iloc[0]
            month_name = first_date.strftime('%B')  # 例如 "March"
            year = str(first_date.year)  # 例如 "2026"
            final_filename = f"{month_name} {year} Turnover.xlsx"
            new_sheet_name = f"{month_name} {year} Turnover"
            print(f"    - 从 Posting Date 提取到日期: {first_date.strftime('%d/%m/%Y')}")
            print(f"    - 最终文件名将为: {final_filename}")
            print(f"    - 'Final Working' 工作表将重命名为: {new_sheet_name}")
        else:
            print("    - 警告: COA 中 'Posting Date' 列没有有效的日期数据，将使用默认名称 'Turnover Report.xlsx'")
            final_filename = "Turnover Report.xlsx"
            new_sheet_name = "Final Working"
    else:
        print(f"    - 警告: COA 文件中未找到 'Posting Date' 列，将使用默认名称 'Turnover Report.xlsx'")
        final_filename = "Turnover Report.xlsx"
        new_sheet_name = "Final Working"
    # --- 新增步骤结束 ---

    df_coa_processed = preprocess_coa(df_coa_raw)

    df_template = pd.read_excel(TEMPLATE_FILE, sheet_name='Final working')
    print(f"    - 成功加载 Final working 模板结构，形状: {df_template.shape}")

    # 创建映射后的输出表
    df_output = pd.DataFrame(index=df_coa_processed.index)
    for template_col in df_template.columns:
        if template_col in COA_TO_FINAL_WORKING_MAP.values():
            coa_col = next((k for k, v in COA_TO_FINAL_WORKING_MAP.items() if v == template_col), None)
            if coa_col and coa_col in df_coa_processed.columns:
                df_output[template_col] = df_coa_processed[coa_col]
            else:
                df_output[template_col] = pd.NA
        else:
            df_output[template_col] = pd.NA

    print(f"    - 映射完成，输出表形状: {df_output.shape}")

except FileNotFoundError as e:
    print(f"    - 错误: 找不到文件 - {e}")
    df_output = pd.DataFrame()
    final_filename = "Turnover Report.xlsx"
    new_sheet_name = "Final Working"
except Exception as e:
    print(f"    - 处理COA时发生错误: {e}")
    df_output = pd.DataFrame()
    final_filename = "Turnover Report.xlsx"
    new_sheet_name = "Final Working"

# --- 步骤 5: 保存到集成调试文件 ---
print("\n--- 步骤 5: 保存到集成调试文件 ---")
integration_debug_file = "full_integration_debug.xlsx"

with pd.ExcelWriter(integration_debug_file, engine='openpyxl') as writer:
    # 写入各个DataFrame
    if not df_inv_cn.empty:
        df_inv_cn.to_excel(writer, sheet_name='INV & CN', index=False)
    if not df_line.empty:
        df_line.to_excel(writer, sheet_name='Inv Line', index=False)
    if not df_cust_industry.empty:
        df_cust_industry.to_excel(writer, sheet_name='Cust-Indust-M', index=False)
    if not df_output.empty:
        df_output.to_excel(writer, sheet_name='Final Working', index=False)

# --- 步骤 6: 格式化 Final Working 工作表 ---
print("\n--- 步骤 6: 格式化 Final Working 工作表 ---")
wb = load_workbook(integration_debug_file)
ws_final = wb['Final Working']

# 1. 获取列名列表，用于定位特定列
column_names = [cell.value for cell in ws_final[1]]

# 2. 获取数据总行数，用于后续操作
total_data_rows = len(df_output) + 1  # +1 因为有标题行

# 3. 执行VLOOKUP并直接写入值
# 创建辅助字典以加速查找
# Inv Line: A列 -> B列
df_line_dict = {}
if not df_line.empty and len(df_line.columns) > 1:
    df_line_dict = dict(zip(df_line.iloc[:, 0], df_line.iloc[:, 1]))

# INV & CN: A列 -> [B, C, D, E] 列
df_inv_cn_dict = {}
if not df_inv_cn.empty and len(df_inv_cn.columns) > 4:
    df_inv_cn_dict = df_inv_cn.set_index(df_inv_cn.columns[0]).T.to_dict('dict')

# Cust-Indust-M: A列 -> B列
df_cust_industry_dict = {}
if not df_cust_industry.empty and len(df_cust_industry.columns) > 1:
    df_cust_industry_dict = dict(zip(df_cust_industry.iloc[:, 0], df_cust_industry.iloc[:, 1]))

for row_num in range(2, total_data_rows + 1):
    # Column F: VLOOKUP from 'Inv Line' (C列 -> Inv Line A列)
    lookup_val_f = ws_final.cell(row=row_num, column=3).value # C列的值
    result_f = df_line_dict.get(lookup_val_f, None)
    ws_final.cell(row=row_num, column=6, value=result_f) # F列

    # Column G: VLOOKUP from 'INV & CN' (C列 -> INV & CN A列) -> B列
    lookup_val_g = ws_final.cell(row=row_num, column=3).value
    result_g_dict = df_inv_cn_dict.get(lookup_val_g, {})
    result_g = result_g_dict.get(df_inv_cn.columns[1], None) if isinstance(result_g_dict, dict) else None
    ws_final.cell(row=row_num, column=7, value=result_g) # G列

    # Column H: VLOOKUP from 'INV & CN' (C列 -> INV & CN A列) -> C列
    lookup_val_h = ws_final.cell(row=row_num, column=3).value
    result_h_dict = df_inv_cn_dict.get(lookup_val_h, {})
    result_h = result_h_dict.get(df_inv_cn.columns[2], None) if isinstance(result_h_dict, dict) else None
    ws_final.cell(row=row_num, column=8, value=result_h) # H列

    # Column I: VLOOKUP from 'INV & CN' (C列 -> INV & CN A列) -> D列
    lookup_val_i = ws_final.cell(row=row_num, column=3).value
    result_i_dict = df_inv_cn_dict.get(lookup_val_i, {})
    result_i = result_i_dict.get(df_inv_cn.columns[3], None) if isinstance(result_i_dict, dict) else None
    ws_final.cell(row=row_num, column=9, value=result_i) # I列

    # Column J: VLOOKUP from 'INV & CN' (C列 -> INV & CN A列) -> E列
    lookup_val_j = ws_final.cell(row=row_num, column=3).value
    result_j_dict = df_inv_cn_dict.get(lookup_val_j, {})
    result_j = result_j_dict.get(df_inv_cn.columns[4], None) if isinstance(result_j_dict, dict) else None
    ws_final.cell(row=row_num, column=10, value=result_j) # J列

    # Column K: VLOOKUP from 'Cust-Indust-M' (H列 -> Cust-Indust-M A列) -> B列
    lookup_val_k = ws_final.cell(row=row_num, column=8).value # H列的值
    result_k = df_cust_industry_dict.get(lookup_val_k, None)
    ws_final.cell(row=row_num, column=11, value=result_k) # K列


# 4. 将 F 列到 K 列 (索引 6-11) 的空单元格填入 "N/A"
for row_num in range(2, total_data_rows + 1):
    for col_num in range(6, 12): # F(6) 到 K(11)
        cell = ws_final.cell(row=row_num, column=col_num)
        if cell.value is None or cell.value == "":
            cell.value = "N/A"

# 5. 处理 Document Type 列的空白单元格
doc_type_col_idx = column_names.index('Document Type') + 1
for row_num in range(2, total_data_rows + 1):
    doc_type_cell = ws_final.cell(row=row_num, column=doc_type_col_idx)
    if doc_type_cell.value is None or doc_type_cell.value == "":
        doc_type_cell.value = "N/A"

# 6. 格式化货币列 (Debit Amount, Credit Amount, Total Amount) 并处理Total Amount
name_col_idx = column_names.index('Name') + 1  # openpyxl 是基于1的索引
debit_col_idx = column_names.index('Debit Amount') + 1
credit_col_idx = column_names.index('Credit Amount') + 1
# 尝试查找 Total Amount 列，如果不存在则创建
try:
    total_amount_col_idx = column_names.index('Total Amount') + 1
except ValueError:
    # 如果 'Total Amount' 列不存在，则添加它（假设在Credit Amount之后）
    total_amount_col_idx = credit_col_idx + 1
    ws_final.cell(row=1, column=total_amount_col_idx, value='Total Amount')
    print(f"    - 已添加 'Total Amount' 列 (列 {ws_final.cell(row=1, column=total_amount_col_idx).column_letter})")

# 7. 设置 Name 列的宽度
ws_final.column_dimensions[ws_final.cell(row=1, column=name_col_idx).column_letter].width = 30

for row_num in range(2, total_data_rows + 1): # 对所有数据行进行格式化
    # 格式化 Debit Amount
    debit_cell = ws_final.cell(row=row_num, column=debit_col_idx)
    debit_cell.number_format = '$#,##0.00'
    if debit_cell.value is not None and isinstance(debit_cell.value, (int, float)) and debit_cell.value < 0:
        debit_cell.font = Font(color="FF0000", bold=False) # 红色字体
    
    # 格式化 Credit Amount
    credit_cell = ws_final.cell(row=row_num, column=credit_col_idx)
    credit_cell.number_format = '$#,##0.00'
    
    # 计算 Total Amount 列的值并格式化
    debit_val = ws_final.cell(row=row_num, column=debit_col_idx).value or 0
    credit_val = ws_final.cell(row=row_num, column=credit_col_idx).value or 0
    total_val = debit_val + credit_val if pd.notna(debit_val) and pd.notna(credit_val) else ""
    
    total_cell = ws_final.cell(row=row_num, column=total_amount_col_idx)
    total_cell.value = total_val
    total_cell.number_format = '$#,##0.00'


# 8. 添加底部汇总行 (现在是 total_data_rows + 1)
summary_row = total_data_rows + 1

# 将 'Total' 文字放在 K 列
ws_final.cell(row=summary_row, column=11, value='Total').font = Font(bold=True) # K列是第11列

# 计算并设置汇总值
sum_debit = sum(
    ws_final.cell(row=r, column=debit_col_idx).value 
    for r in range(2, total_data_rows + 1) 
    if isinstance(ws_final.cell(row=r, column=debit_col_idx).value, (int, float))
)
sum_credit = sum(
    ws_final.cell(row=r, column=credit_col_idx).value 
    for r in range(2, total_data_rows + 1) 
    if isinstance(ws_final.cell(row=r, column=credit_col_idx).value, (int, float))
)

sum_debit_cell = ws_final.cell(row=summary_row, column=debit_col_idx)
sum_debit_cell.value = sum_debit
sum_debit_cell.font = Font(bold=True, color="FF0000")
sum_debit_cell.number_format = '$#,##0.00'

sum_credit_cell = ws_final.cell(row=summary_row, column=credit_col_idx)
sum_credit_cell.value = sum_credit
sum_credit_cell.font = Font(bold=True)
sum_credit_cell.number_format = '$#,##0.00'

# 为汇总行的 Total Amount 列写入值
sum_total_cell = ws_final.cell(row=summary_row, column=total_amount_col_idx)
sum_total_cell.value = sum_debit + sum_credit
sum_total_cell.font = Font(bold=True) # 默认为加粗
sum_total_cell.number_format = '$#,##0.00'

# 9. 为标题行设置筛选器，并将内容单元格背景设置为无色
last_col_letter = ws_final.cell(row=1, column=len(column_names)).column_letter
data_range = f"A1:{last_col_letter}{total_data_rows}" # 定义包含标题和数据的范围
tab = Table(displayName="FinalWorkingTable", ref=data_range) # 创建一个表格对象

# 定义表格样式，不使用任何内置样式，以确保背景为白色/无色
style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
tab.tableStyleInfo = style

ws_final.add_table(tab) # 将表格（含筛选器）添加到工作表

# 10. 格式化标题行 (第1行) 并冻结窗格
title_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # white, Background 1, Darker 25%
for col_num in range(1, len(column_names) + 1):
    cell = ws_final.cell(row=1, column=col_num)
    cell.fill = title_fill
ws_final.freeze_panes = "A2"  # 冻结首行

# 11. 统一处理 Total Amount 列的负数标红逻辑 (包括数据行和汇总行)
for row_num in range(2, summary_row + 1): # 从数据第一行处理到汇总行
    total_cell = ws_final.cell(row=row_num, column=total_amount_col_idx)
    # 检查单元格的值是否为负数
    if total_cell.value is not None:
        try:
            # 如果值是数值类型且小于0
            if total_cell.value < 0:
                # 如果是负数，且当前字体不是粗体（说明不是汇总行），则只标红
                if not total_cell.font.bold:
                    total_cell.font = Font(color="FF0000", bold=False)
                # 如果是负数，且当前字体是粗体（说明是汇总行），则保持粗体并标红
                else:
                    total_cell.font = Font(bold=True, color="FF0000")
        except TypeError:
            # 如果 total_cell.value 不是数值（例如，公式本身返回了文本），则跳过
            continue


# --- 步骤 7: 创建 Top 10 Salesperson 工作表 ---
print("\n--- 步骤 7: 创建 Top 10 Salesperson 工作表 ---")

# 1. 读取Final Working工作表的数据到DataFrame
data = ws_final.values
columns = next(data) # 获取列名
df_for_chart = pd.DataFrame(data, columns=columns)

# 2. 检查是否存在'Salesperson Code'列
if 'Salesperson Code' in df_for_chart.columns:
    print("    - 发现 'Salesperson Code' 列，开始处理数据...")
    # 3. 按销售员代码分组并聚合数据
    sales_data = df_for_chart.groupby('Salesperson Code').agg({
        'Debit Amount': 'sum',
        'Credit Amount': 'sum'
    }).reset_index()

    # 4. 计算Total Amount
    sales_data['Total Amount'] = sales_data['Debit Amount'] + sales_data['Credit Amount']

    # 5. 按Total Amount降序排列并选择前10名
    top_sales_data = sales_data.sort_values(by='Total Amount', ascending=False).head(10)

    # 6. 创建新的工作表"Top 10 Salesperson"
    if 'Top 10 Salesperson' in wb.sheetnames:
        del wb['Top 10 Salesperson']  # 如果已存在则删除旧的工作表
    ws_top_sales = wb.create_sheet(title="Top 10 Salesperson")

    # 7. 将标题写入新工作表
    titles = ['Salesperson Code', 'Debit Amount', 'Credit Amount', 'Total Amount']
    for col_num, title in enumerate(titles, 1):
        ws_top_sales.cell(row=1, column=col_num, value=title)

    # 8. 写入数据
    for row_num, row_data in enumerate(top_sales_data.itertuples(), 2):
        for col_num, value in enumerate(row_data[1:], 1):  # 忽略itertuples返回的第一个索引值
            ws_top_sales.cell(row=row_num, column=col_num, value=value)

    # 9. 格式化金额列 (Debit Amount, Credit Amount, Total Amount)
    # 定位各金额列的索引
    debit_sales_col_idx = titles.index('Debit Amount') + 1
    credit_sales_col_idx = titles.index('Credit Amount') + 1
    total_sales_col_idx = titles.index('Total Amount') + 1

    for row_num in range(2, len(top_sales_data) + 2): # 从数据第一行开始
        # 格式化 Debit Amount
        debit_cell = ws_top_sales.cell(row=row_num, column=debit_sales_col_idx)
        debit_cell.number_format = '$#,##0.00'
        if debit_cell.value is not None and isinstance(debit_cell.value, (int, float)) and debit_cell.value < 0:
            debit_cell.font = Font(color="FF0000") # 红色字体

        # 格式化 Credit Amount
        credit_cell = ws_top_sales.cell(row=row_num, column=credit_sales_col_idx)
        credit_cell.number_format = '$#,##0.00'
        if credit_cell.value is not None and isinstance(credit_cell.value, (int, float)) and credit_cell.value < 0:
            credit_cell.font = Font(color="FF0000") # 红色字体

        # 格式化 Total Amount
        total_cell = ws_top_sales.cell(row=row_num, column=total_sales_col_idx)
        total_cell.number_format = '$#,##0.00'
        if total_cell.value is not None and isinstance(total_cell.value, (int, float)) and total_cell.value < 0:
            total_cell.font = Font(color="FF0000") # 红色字体


    # 10. 创建柱状图
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Top 10 Salesperson Performance"
    chart.y_axis.title = 'Amount'
    chart.x_axis.title = 'Salesperson Code'

    # 11. 设置图表数据范围
    data = Reference(ws_top_sales, min_col=2, max_col=4, min_row=1, max_row=top_sales_data.shape[0] + 1)
    categories = Reference(ws_top_sales, min_col=1, min_row=2, max_row=top_sales_data.shape[0] + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.shape = 4

    # 12. 添加图表到工作表
    ws_top_sales.add_chart(chart, "F5")  # 在指定位置添加图表
    print("    - 'Top 10 Salesperson' 工作表及柱状图已成功生成。")
else:
    print("    - 警告: 'Final Working' 工作表中未找到 'Salesperson Code' 列，无法生成 'Top 10 Salesperson' 工作表。")


# --- 步骤 8: 重命名Final Working工作表 ---
print("\n--- 步骤 8: 重命名 'Final Working' 工作表 ---")
if 'Final Working' in wb.sheetnames:
    wb['Final Working'].title = new_sheet_name
    print(f"    - 'Final Working' 工作表已重命名为: {new_sheet_name}")
else:
    print("    - 警告: 未找到 'Final Working' 工作表。")

# --- 步骤 9: 只保留指定工作表 ---
print("\n--- 步骤 9: 清理工作表，只保留最终所需的 ---")
sheets_to_keep = [new_sheet_name, 'Top 10 Salesperson']
all_sheets = list(wb.sheetnames)

for sheet_name in all_sheets:
    if sheet_name not in sheets_to_keep:
        del wb[sheet_name]
        print(f"    - 已删除工作表: {sheet_name}")

print(f"    - 最终保留的工作表: {list(wb.sheetnames)}")

# --- 步骤 10: 自动调整所有最终工作表的列宽 ---
print("\n--- 步骤 10: 自动调整最终工作表的列宽 ---")
for sheet_name in wb.sheetnames:
    worksheet = wb[sheet_name]
    print(f"    - 正在调整工作表 '{sheet_name}' 的列宽...")
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # 获取列字母 (e.g., 'A', 'B')
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 增加2个单位作为边距
        worksheet.column_dimensions[column].width = adjusted_width if adjusted_width > 10 else 10 # 设置最小宽度为10
    print(f"    - 工作表 '{sheet_name}' 的列宽调整完成。")

# --- 步骤 11: 保存最终文件 ---
print(f"\n--- 步骤 11: 保存最终文件 ---")
# 如果最终文件名与临时文件名相同，为了避免冲突，先保存为临时文件再移动
if final_filename == integration_debug_file:
    temp_name = "temp_output_" + final_filename
    wb.save(temp_name)
    if os.path.exists(final_filename):
        os.remove(final_filename)
    os.rename(temp_name, final_filename)
    print(f"    - 已保存最终文件为: {final_filename}")
else:
    wb.save(final_filename)
    print(f"    - 已保存最终文件为: {final_filename}")

# 删除中间调试文件
if os.path.exists(integration_debug_file):
    os.remove(integration_debug_file)
    print(f"    - 已删除中间调试文件: {integration_debug_file}")

print(f"\n所有步骤已完成。最终结果已保存至: {final_filename}")
print("- Sheet 1 (可能已重命名): 包含完整的账目流水数据，已完成所有格式化。")
print("- Sheet 2 ('Top 10 Salesperson'): 包含按业绩排序的前10名销售员数据及柱状图。")
print("   - 金额列已格式化为货币($)，负数已标红。")
print("   - 所有工作表的列宽均已自动调整以适应内容。")