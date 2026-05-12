import os
import re
import logging
import base64
import requests
import msal
import pytz
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================== 多租户配置区域 ==================
MAILBOX_CONFIGS = [
    {
        "email": "sample_1@email.com",
        "tenant_id": "sample_tenant_id_1",
        "client_id": "sample_client_id_1",
        "client_secret": "sample_secret_1"
    },
    {
        "email": "sample_2@email.com",
        "tenant_id": "sample_tenant_id_2",
        "client_id": "sample_client_id_2",
        "client_secret": "sample_secret_2"
    }
]

SENDER_EMAIL = "sample_1@email.com"
REPORT_RECIPIENT = "hubin@jpnelson.com.sg"
CC_RECIPIENT = "hubin@jpnelson.com.sg"

TARGET_RECIPIENTS = {
    'sample_1@email.com', 'sample_2@email.com', 'sample_3@email.com',
    'User 1', 'User 2', 'User 3'
}
# ===================================================

SG_TZ = pytz.timezone('Asia/Singapore')

def setup_logging():
    """配置日志：适配 Linux 路径，仅输出到文件"""
    # 使用 os.path.join 确保在 Linux 下生成 Logs/Weekly_Report_Log
    logs_dir = os.path.join("Logs", "Weekly_Report_Log")
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir, exist_ok=True)
    
    date_str = datetime.now(SG_TZ).strftime('%d_%m_%Y')
    log_file = os.path.join(logs_dir, f"Weekly_Enquiry_Summary_log_{date_str}.log")
    
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - [%(levelname)s] - %(message)s',
        handlers=[logging.FileHandler(log_file, encoding='utf-8')]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

def normalize_subject_v3(subject):
    if not subject: return ""
    sub = re.sub(r'^(FW|FWD|RE|答复|转发|External|Automatic reply):\s*', '', subject.strip(), flags=re.IGNORECASE)
    sub = re.sub(r'\[.*?\]', '', sub)
    return ' '.join(sub.lower().split()).strip()

def get_token(config):
    logger.info(f"Action: Requesting Token for {config['email']}")
    authority = f"https://login.microsoftonline.com/{config['tenant_id']}"
    app = msal.ConfidentialClientApplication(config['client_id'], authority=authority, client_credential=config['client_secret'])
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" in result:
        return result["access_token"]
    logger.error(f"Error: Token failed for {config['email']}: {result.get('error')}")
    return None

def fetch_all_messages_cache(token, mailbox, start_date):
    """搜索全邮箱消息，适配 Linux 环境下的 API 调用"""
    logger.info(f"Step: Fetching all messages for {mailbox}")
    cache = {}
    url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/messages"
    headers = {"Authorization": f"Bearer {token}"}
    start_utc = start_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    params = {
        "$select": "subject,receivedDateTime,sender,toRecipients",
        "$filter": f"receivedDateTime ge {start_utc}",
        "$orderby": "receivedDateTime desc",
        "$top": 999
    }

    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code != 200: break
        data = resp.json()
        for item in data.get('value', []):
            received_time = item.get('receivedDateTime', '')
            clean_sub = normalize_subject_v3(item.get('subject', ''))
            sender = (item.get('sender') or {}).get('emailAddress', {}).get('address', '')
            to_list = [r.get('emailAddress', {}).get('address', '') for r in item.get('toRecipients', [])]
            dt = datetime.strptime(received_time[:19], '%Y-%m-%dT%H:%M:%S').replace(tzinfo=pytz.utc).astimezone(SG_TZ)
            cache[clean_sub] = {
                'received_time': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'sender': sender,
                'to': "; ".join(filter(None, to_list)),
                'full_subject': item.get('subject', '')
            }
        url = data.get('@odata.nextLink')
        params = None
    return cache

def fetch_sent_forwards(token, mailbox, start_date):
    logger.info(f"Step: Fetching SentItems for {mailbox}")
    forwards = []
    url = f"https://graph.microsoft.com/v1.0/users/{mailbox}/mailFolders/SentItems/messages"
    headers = {"Authorization": f"Bearer {token}"}
    start_utc = start_date.astimezone(pytz.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    params = {
        "$select": "subject,sentDateTime,toRecipients",
        "$filter": f"sentDateTime ge {start_utc}",
        "$orderby": "sentDateTime desc"
    }

    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code != 200: break
        data = resp.json()
        for item in data.get('value', []):
            subject = item.get('subject', '')
            if not any(p in subject.upper() for p in ['FW:', 'FWD:']): continue
            matches = []
            for r in item.get('toRecipients', []):
                name, addr = r['emailAddress'].get('name', ''), r['emailAddress'].get('address', '')
                if any(t.lower() in (name+addr).lower() for t in TARGET_RECIPIENTS):
                    matches.append(name or addr)
            if matches:
                dt = datetime.strptime(item['sentDateTime'][:19], '%Y-%m-%dT%H:%M:%S').replace(tzinfo=pytz.utc).astimezone(SG_TZ)
                forwards.append({
                    'raw_subject': subject,
                    'clean_subject': normalize_subject_v3(subject),
                    'forwarded_to': ", ".join(sorted(set(matches))),
                    'forwarded_on': dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'sort_key': dt
                })
        url = data.get('@odata.nextLink')
        params = None
    return forwards

def main():
    logger.info("--- Linux Adaption Version Started ---")
    now_sg = datetime.now(SG_TZ)
    monday = (now_sg - timedelta(days=now_sg.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    deep_history = (now_sg.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    combined_data = []
    sender_token = None

    for config in MAILBOX_CONFIGS:
        logger.info(f"Action: Processing {config['email']}")
        token = get_token(config)
        if not token: continue
        if config['email'] == SENDER_EMAIL: sender_token = token
            
        all_msg_cache = fetch_all_messages_cache(token, config['email'], deep_history)
        forwards = fetch_sent_forwards(token, config['email'], monday)
        
        for f in forwards:
            match = all_msg_cache.get(f['clean_subject'])
            if not match:
                for k, v in all_msg_cache.items():
                    if f['clean_subject'] and (f['clean_subject'] in k or k in f['clean_subject']):
                        match = v; break
            combined_data.append({
                'Source Mailbox': config['email'],
                'Sent on': match['received_time'] if match else '',
                'Sender Address': match['sender'] if match else '',
                'Sent to': match['to'] if match else '',
                'Forwarded to': f['forwarded_to'],
                'Forwarded on': f['forwarded_on'],
                'Title': match['full_subject'] if match else f['raw_subject'],
                '_sort_date': f['sort_key']
            })

    if not combined_data:
        logger.warning("End: No records found.")
        return

    df = pd.DataFrame(combined_data)
    df = df.sort_values(by='_sort_date', ascending=False).drop(columns=['_sort_date'])

    date_str = now_sg.strftime('%d_%m_%Y')
    report_name = f"Weekly Enquiry Summary {date_str}.xlsx"
    
    # 适配 Linux 路径
    output_dir = "Weekly_Reports"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, report_name)
    
    df.to_excel(output_path, index=False)
    
    # 格式美化
    wb = load_workbook(output_path)
    ws = wb.active
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    wb.save(output_path)

    # 发送邮件
    if sender_token:
        try:
            with open(output_path, "rb") as f:
                content = base64.b64encode(f.read()).decode()
            payload = {
                "message": {
                    "subject": f"Weekly Enquiry Summary {now_sg.strftime('%Y%m%d')}",
                    "body": {"contentType": "Text", "content": "Hi Mr Nelson,\n\nPlease find the attached report, thanks."},
                    "toRecipients": [{"emailAddress": {"address": REPORT_RECIPIENT}}],
                    "ccRecipients": [{"emailAddress": {"address": CC_RECIPIENT}}],
                    "attachments": [{
                        "@odata.type": "#microsoft.graph.fileAttachment",
                        "name": report_name,
                        "contentBytes": content
                    }]
                }
            }
            requests.post(f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail", 
                          headers={"Authorization": f"Bearer {sender_token}"}, json=payload)
            logger.info("Success: Email sent.")
        except Exception as e:
            logger.error(f"Error during send: {e}")

    logger.info("--- Execution Completed ---")

if __name__ == "__main__":
    main()